"""Check the actual column layout of the Evaluations tab."""
import requests, io
from openpyxl import load_workbook

sheet_id = "1J-pZGelB9DxtahUc1JL3IXkT5C2_ajd_qvE_oqxUia4"
xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
resp = requests.get(xlsx_url, timeout=30)
wb = load_workbook(io.BytesIO(resp.content), data_only=True)

ws = wb['Evaluations']
# Print all headers (row 1) with column letters
from openpyxl.utils import get_column_letter
print("--- Evaluations Tab Headers ---")
for col in range(1, ws.max_column + 1):
    letter = get_column_letter(col)
    val = ws.cell(row=1, column=col).value
    if val:
        print(f"  {letter:4s} ({col:2d}): {val}")

# Now check what columns O and AB contain
print(f"\n--- Column O header: {ws.cell(row=1, column=15).value} ---")
print(f"--- Column AB header: {ws.cell(row=1, column=28).value} ---")

# Sample values from columns O and AB (first 5 data rows)
print("\n--- Sample Column O values ---")
for r in range(2, 8):
    print(f"  Row {r}: {ws.cell(row=r, column=15).value}")

print("\n--- Sample Column AB values ---")
for r in range(2, 8):
    print(f"  Row {r}: {ws.cell(row=r, column=28).value}")

# Count non-empty numeric values in O and AB
o_count = 0
o_sum = 0.0
ab_count = 0
ab_sum = 0.0
for r in range(2, ws.max_row + 1):
    ov = ws.cell(row=r, column=15).value
    abv = ws.cell(row=r, column=28).value
    if ov is not None and isinstance(ov, (int, float)):
        o_count += 1
        o_sum += ov
    if abv is not None and isinstance(abv, (int, float)):
        ab_count += 1
        ab_sum += abv

print(f"\n--- Column O: count={o_count}, sum={o_sum:.2f} ---")
print(f"--- Column AB: count={ab_count}, sum={ab_sum:.2f} ---")
print(f"--- EV = (o_sum + ab_sum) / (o_count + ab_count) = {(o_sum + ab_sum) / (o_count + ab_count) if (o_count + ab_count) > 0 else 0:.2f} ---")

wb.close()
