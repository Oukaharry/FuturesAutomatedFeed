"""Fetch Chris's Google Sheet XLSX and compare Hedge Day data with DB."""
import urllib.request, io, json, sqlite3, sys
try:
    import openpyxl
except ImportError:
    print("Need openpyxl")
    sys.exit(1)

sheet_key = '1q4atojmjW03XLU6bRfubZ3WZiK071x3eQttt5kdKVYs'
xlsx_url = f'https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=xlsx'

print("Fetching XLSX from Google Sheets...")
req = urllib.request.Request(xlsx_url, headers={'User-Agent': 'Mozilla/5.0'})
resp = urllib.request.urlopen(req, timeout=60)
data = resp.read()
print(f"Downloaded {len(data)} bytes")

wb = openpyxl.load_workbook(filename=io.BytesIO(data), data_only=True)
print(f"Sheet names: {wb.sheetnames}")

# Find the evaluations tab (first sheet usually)
ws = wb[wb.sheetnames[0]]

# Find header row
header_idx = None
col_map = {}
for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
    row_vals = [str(c).strip() if c else '' for c in row]
    if any('Prop Firm' in v for v in row_vals):
        header_idx = r_idx
        col_map = {}
        for ci, h in enumerate(row_vals):
            if h:
                col_map[ci] = h
        break
    elif any('Account Size' in str(v) for v in row_vals):
        header_idx = r_idx
        col_map = {}
        for ci, h in enumerate(row_vals):
            if h:
                col_map[ci] = h
        if 0 not in col_map or not col_map[0]:
            col_map[0] = 'Prop Firm'
        break

if header_idx is None:
    print("ERROR: Could not find header row!")
    sys.exit(1)

print(f"Header at row {header_idx}")

# Find Hedge Day column indices
hedge_cols = {}
account_col = None
account1_col = None
prop_firm_col = None
status_p1_col = None
status_col = None
status_funded_col = None

for ci, name in col_map.items():
    if name.startswith('Hedge Day '):
        try:
            num = int(name.replace('Hedge Day ', ''))
            hedge_cols[num] = ci
        except ValueError:
            pass
    elif name == 'Account #':
        account_col = ci
    elif name == 'Account #.1':
        account1_col = ci
    elif name == 'Prop Firm':
        prop_firm_col = ci
    elif name == 'Status P1':
        status_p1_col = ci
    elif name == 'Status Funded':
        status_funded_col = ci
    elif name == 'Status' and status_col is None:
        status_col = ci

print(f"Found Hedge Day columns: {sorted(hedge_cols.keys())}")

# Parse all data rows
sheet_evals = []
for row in ws.iter_rows(min_row=header_idx+1, values_only=True):
    row_vals = list(row)
    pf = str(row_vals[prop_firm_col]).strip() if prop_firm_col is not None and row_vals[prop_firm_col] else ''
    if not pf or pf == 'None':
        continue
    
    ev = {
        'Prop Firm': pf,
        'Account #': str(row_vals[account_col]).strip() if account_col is not None and row_vals[account_col] else '',
        'Account #.1': str(row_vals[account1_col]).strip() if account1_col is not None and row_vals[account1_col] else '',
        'Status P1': str(row_vals[status_p1_col]).strip() if status_p1_col is not None and row_vals[status_p1_col] else '',
    }
    
    if status_funded_col is not None and row_vals[status_funded_col]:
        ev['Status Funded'] = str(row_vals[status_funded_col]).strip()
    elif status_col is not None and row_vals[status_col]:
        ev['Status Funded'] = str(row_vals[status_col]).strip()
    
    for d, ci in hedge_cols.items():
        val = row_vals[ci] if ci < len(row_vals) else None
        ev[f'Hedge Day {d}'] = val
    
    sheet_evals.append(ev)

print(f"Total sheet evaluations: {len(sheet_evals)}")

# Now load DB data
conn = sqlite3.connect('dashboard/dashboard.db')
cursor = conn.cursor()
cursor.execute('SELECT evaluations FROM clients_data WHERE client_id = ?', ('Chris',))
row = cursor.fetchone()
db_evals = json.loads(row[0]) if row else []
conn.close()
print(f"Total DB evaluations: {len(db_evals)}")

# Target accounts to compare
target_accounts = {
    '2641': 'V2-2641 (eval 392)',
    '6337': 'V2-6337 (eval 392)',
    '76770': 'FNFT-76770 (eval 445)',
    '46494': 'FNFT-46494 (eval 446)',
    '57582': 'TDF-57582 (eval 448)',
    '33548': 'TDF-33548 (eval 449)',
    '80230': 'MFFU-80230 (UNMATCHED)',
    '80229': 'MFFU-80229 (eval 394 inactive)',
    '80233': 'MFFU-80233 (eval 408 inactive)',
    '80237': 'MFFU-80237 (eval 426 inactive)',
    '59522': 'TDF-59522 (eval 447 inactive)',
    '23825': 'FNFT-23825 (eval 277 inactive)',
    '62524': 'FNFT-62524 (eval 282 inactive)',
}

def normalize_val(v):
    """Normalize a hedge day value to a float for comparison."""
    if v is None or v == '' or v == 'None':
        return None
    s = str(v).replace('$', '').replace(',', '').strip()
    if not s or s == '0' or s == '0.00' or s == '$0.00':
        return 0.0
    try:
        return round(float(s), 2)
    except:
        return None

print("\n" + "="*80)
print("COMPARING SHEET vs DB FOR FA ACCOUNTS:")
print("="*80)

for target, label in sorted(target_accounts.items(), key=lambda x: x[1]):
    # Find in sheet
    sheet_matches = []
    for si, sev in enumerate(sheet_evals):
        acc = sev.get('Account #', '')
        acc1 = sev.get('Account #.1', '')
        if target in acc or target in acc1:
            sheet_matches.append((si, sev))
    
    # Find in DB
    db_matches = []
    for di, dev in enumerate(db_evals):
        acc = str(dev.get('Account #', ''))
        acc1 = str(dev.get('Account #.1', ''))
        if target in acc or target in acc1:
            db_matches.append((di, dev))
    
    if not sheet_matches and not db_matches:
        continue
    
    print(f"\n{'='*70}")
    print(f"  {label}")
    print(f"{'='*70}")
    
    for si, sev in sheet_matches:
        print(f"\n  [SHEET row {si+2}] Account: {sev.get('Account #')} / {sev.get('Account #.1')}")
        print(f"    Prop Firm: {sev.get('Prop Firm')}  Status P1: {sev.get('Status P1')}  Status: {sev.get('Status Funded', '')}")
        
        # Find matching DB eval
        di_match = None
        dev_match = None
        for di, dev in db_matches:
            if di == si:  # Same index
                di_match = di
                dev_match = dev
                break
        
        if dev_match is None and db_matches:
            di_match, dev_match = db_matches[0]
        
        if dev_match:
            print(f"  [DB   idx {di_match} row {di_match+2}] Account: {dev_match.get('Account #')} / {dev_match.get('Account #.1')}")
            print(f"    Prop Firm: {dev_match.get('Prop Firm')}  Status P1: {dev_match.get('Status P1')}  Status: {dev_match.get('Status', dev_match.get('Status Funded', ''))}")
        else:
            print(f"  [DB] NOT FOUND for this account!")
        
        # Compare hedge days
        print(f"\n    {'Hedge Day':<15} {'SHEET':<20} {'DB':<20} {'Match?':<10}")
        print(f"    {'-'*65}")
        
        total_sheet = 0.0
        total_db = 0.0
        mismatches = 0
        
        for d in range(1, 35):
            s_val = sev.get(f'Hedge Day {d}')
            d_val = dev_match.get(f'Hedge Day {d}') if dev_match else None
            
            s_norm = normalize_val(s_val)
            d_norm = normalize_val(d_val)
            
            if s_norm is None and d_norm is None:
                continue
            if s_norm == 0 and d_norm is None:
                continue
            if s_norm is None and d_norm == 0:
                continue
            
            if s_norm: total_sheet += s_norm
            if d_norm: total_db += d_norm
            
            match = '✓' if s_norm == d_norm else '✗ DIFF'
            if s_norm != d_norm:
                mismatches += 1
            
            s_display = f'${s_norm:.2f}' if s_norm is not None else '(empty)'
            d_display = f'${d_norm:.2f}' if d_norm is not None else '(empty)'
            print(f"    Hedge Day {d:<3} {s_display:<20} {d_display:<20} {match}")
        
        print(f"    {'-'*65}")
        print(f"    {'TOTALS':<15} ${total_sheet:<19.2f} ${total_db:<19.2f} {'✓' if abs(total_sheet - total_db) < 0.01 else '✗ DIFF: $' + str(round(total_sheet - total_db, 2))}")
        print(f"    Mismatches: {mismatches}")
    
    if not sheet_matches:
        print(f"  NOT FOUND IN SHEET!")
        for di, dev in db_matches:
            print(f"  [DB idx {di}] Account: {dev.get('Account #')} / {dev.get('Account #.1')}")

print("\n\nDone.")
