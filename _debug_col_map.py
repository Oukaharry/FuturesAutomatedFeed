"""Check XLSX column layout and find hedge value differences."""
import sys, os, io
import requests
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

sheet_key = '1q4atojmjW03XLU6bRfubZ3WZiK071x3eQttt5kdKVYs'

print("Fetching XLSX...")
xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=xlsx"
resp = requests.get(xlsx_url, timeout=60)
wb = openpyxl.load_workbook(filename=io.BytesIO(resp.content), data_only=True)
ws = wb[wb.sheetnames[0]]

# Find header and print ALL columns
header_idx = None
for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
    row_vals = [str(c).strip() if c else '' for c in row]
    if any('Prop Firm' in v for v in row_vals):
        header_idx = r_idx
        print(f"Header at row {r_idx}")
        for ci, h in enumerate(row_vals):
            if h:
                col_letter = chr(65 + ci) if ci < 26 else chr(64 + ci // 26) + chr(65 + ci % 26)
                print(f"  Col {col_letter} (idx {ci}): {h}")
        break
