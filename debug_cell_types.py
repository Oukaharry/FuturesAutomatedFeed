"""Check for cells that are TEXT in Google Sheets (SUM ignores) but look numeric in CSV.
Read XLSX to check cell types, compare with CSV-parsed values."""
import requests, io, sys, json, sqlite3
import openpyxl
from decimal import Decimal, InvalidOperation

SHEET_ID = "1EO6-a_b9uun2vwETWu8aGh67ya3nwpdLAo4F-yjc1ZI"
sys.path.insert(0, '.')
from utils.data_processor import parse_currency

# Fetch XLSX
print("Fetching XLSX...")
xlsx_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
resp = requests.get(xlsx_url, timeout=60)
wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
ws = wb['Evaluations']

# Also load without data_only to check cell types
wb_formulas = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)
ws_formulas = wb_formulas['Evaluations']

# Header at row 2
# Hedge columns by letter: J-N (P1), U-AA (funded)
# Payout columns by letter: AC, AE, AG, AI
hedge_cols = {'J': 10, 'K': 11, 'L': 12, 'M': 13, 'N': 14,  # P1 HR1-HR5
              'U': 21, 'V': 22, 'W': 23, 'X': 24, 'Y': 25, 'Z': 26, 'AA': 27}  # Funded HR1.1-HR7
payout_cols = {'AC': 29, 'AE': 31, 'AG': 33, 'AI': 35}  # Payout 1-4

# Get stored evaluations for comparison
conn = sqlite3.connect('dashboard/dashboard.db')
conn.row_factory = sqlite3.Row
cur = conn.cursor()
cur.execute("SELECT evaluations FROM clients_data WHERE client_id='Tyler'")
data = cur.fetchone()
conn.close()
stored = json.loads(data['evaluations'])

# Check cell types and find text-formatted numbers
print(f"\n=== CELLS WITH TEXT-FORMATTED NUMERIC VALUES (SUM ignores these) ===")
text_numeric_hedge = Decimal('0')
text_numeric_payout = Decimal('0')

# Also check for cells that are numeric in XLSX but parse to different values in CSV
data_start = 3  # Row 3 is first data row (after header at row 2)
row_count = 0

for row_idx in range(data_start, ws.max_row + 1):
    # Check if row has Prop Firm
    pf = ws.cell(row=row_idx, column=1).value
    if not pf or not str(pf).strip():
        continue
    
    if row_count >= len(stored):
        break
    
    # Check hedge columns
    for letter, col_num in hedge_cols.items():
        cell = ws.cell(row=row_idx, column=col_num)
        cell_f = ws_formulas.cell(row=row_idx, column=col_num)
        val = cell.value
        val_f = cell_f.value  # Formula or raw value
        
        # Check if the cell is text-formatted but looks like a number
        if val is not None and isinstance(val, str):
            # This cell is stored as text in the XLSX
            s = str(val).strip().replace('$', '').replace(',', '')
            if s and s != '-' and s != 'nan':
                try:
                    num_val = Decimal(s)
                    if num_val != 0:
                        header = ws.cell(row=2, column=col_num).value
                        print(f"  HEDGE TEXT-NUM: Row {row_count}, {header} ({letter}): text='{val}', parsed={num_val}")
                        text_numeric_hedge += num_val
                except:
                    pass
        
        # Also check: is the value a formula that might differ?
        if val_f is not None and isinstance(val_f, str) and val_f.startswith('='):
            header = ws.cell(row=2, column=col_num).value
            print(f"  HEDGE FORMULA: Row {row_count}, {header} ({letter}): formula='{val_f}', cached={val}")
    
    # Check payout columns
    for letter, col_num in payout_cols.items():
        cell = ws.cell(row=row_idx, column=col_num)
        cell_f = ws_formulas.cell(row=row_idx, column=col_num)
        val = cell.value
        val_f = cell_f.value
        
        if val is not None and isinstance(val, str):
            s = str(val).strip().replace('$', '').replace(',', '')
            if s and s != '-' and s != 'nan':
                try:
                    num_val = Decimal(s)
                    if num_val != 0:
                        header = ws.cell(row=2, column=col_num).value
                        print(f"  PAYOUT TEXT-NUM: Row {row_count}, {header} ({letter}): text='{val}', parsed={num_val}")
                        text_numeric_payout += num_val
                except:
                    pass
        
        if val_f is not None and isinstance(val_f, str) and val_f.startswith('='):
            header = ws.cell(row=2, column=col_num).value
            print(f"  PAYOUT FORMULA: Row {row_count}, {header} ({letter}): formula='{val_f}', cached={val}")
    
    row_count += 1

print(f"\n=== SUMMARY ===")
print(f"  Text-formatted numeric hedge values total: {text_numeric_hedge}")
print(f"  Text-formatted numeric payout values total: {text_numeric_payout}")
print(f"  Hedge diff needed: +1.69 (if we subtract text-nums: -26646.11 - ({text_numeric_hedge}) = {Decimal('-26646.11') - text_numeric_hedge})")
print(f"  Payout diff needed: -0.42 (if we subtract text-nums: 145295.62 - ({text_numeric_payout}) = {Decimal('145295.62') - text_numeric_payout})")
print(f"  Stats hedge:  -26644.42")
print(f"  Stats payout: 145295.20")
