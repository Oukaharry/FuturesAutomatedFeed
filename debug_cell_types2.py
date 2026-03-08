"""Simplified: check cell types in XLSX for hedge/payout columns."""
import requests, io, sys
import openpyxl
from decimal import Decimal

SHEET_ID = "1EO6-a_b9uun2vwETWu8aGh67ya3nwpdLAo4F-yjc1ZI"
sys.path.insert(0, '.')

print("Fetching XLSX...", flush=True)
xlsx_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
resp = requests.get(xlsx_url, timeout=60)
print(f"Downloaded {len(resp.content)} bytes", flush=True)

print("Loading workbook...", flush=True)
wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=False)
ws = wb['Evaluations']
print(f"Sheet: {ws.max_row} rows x {ws.max_column} cols", flush=True)

# Check hedge columns J-N and U-AA for text-formatted numbers
# And payout columns AC, AE, AG, AI
hedge_cols = [10,11,12,13,14, 21,22,23,24,25,26,27]  # J-N, U-AA
payout_cols = [29,31,33,35]  # AC, AE, AG, AI

text_hedge = Decimal('0')
text_payout = Decimal('0')
formula_cells = []

print("\nScanning cells...", flush=True)
for row_idx in range(3, ws.max_row + 1):
    pf = ws.cell(row=row_idx, column=1).value
    if not pf or not str(pf).strip():
        continue
    
    for ci in hedge_cols:
        cell = ws.cell(row=row_idx, column=ci)
        val = cell.value
        if val is None:
            continue
        
        # Check if it's a formula
        if isinstance(val, str) and val.startswith('='):
            formula_cells.append((row_idx, ci, val))
            continue
            
        # Check if it's a string that looks numeric
        if isinstance(val, str):
            s = val.strip().replace('$','').replace(',','')
            if s and s != '-' and s != 'nan':
                try:
                    d = Decimal(s)
                    if d != 0:
                        header = ws.cell(row=2, column=ci).value
                        print(f"  TEXT-NUM hedge: R{row_idx} {header}: '{val}' = {d}", flush=True)
                        text_hedge += d
                except:
                    pass
    
    for ci in payout_cols:
        cell = ws.cell(row=row_idx, column=ci)
        val = cell.value
        if val is None:
            continue
        if isinstance(val, str) and val.startswith('='):
            formula_cells.append((row_idx, ci, val))
            continue
        if isinstance(val, str):
            s = val.strip().replace('$','').replace(',','')
            if s and s != '-' and s != 'nan':
                try:
                    d = Decimal(s)
                    if d != 0:
                        header = ws.cell(row=2, column=ci).value
                        print(f"  TEXT-NUM payout: R{row_idx} {header}: '{val}' = {d}", flush=True)
                        text_payout += d
                except:
                    pass

print(f"\nText-formatted numeric totals:", flush=True)
print(f"  Hedge: {text_hedge}", flush=True)
print(f"  Payout: {text_payout}", flush=True)

if formula_cells:
    print(f"\nFormula cells found: {len(formula_cells)}", flush=True)
    for r, c, f in formula_cells[:20]:
        header = ws.cell(row=2, column=c).value
        print(f"  R{r} {header}: {f}", flush=True)

print(f"\nTarget: hedge diff = {Decimal('-26646.11') - text_hedge} (should be -26644.42)", flush=True)
print(f"Target: payout diff = {Decimal('145295.62') - text_payout} (should be 145295.20)", flush=True)
