"""Compute EXACTLY what SUM(I:M) + SUM(T:Z) and SUM(AB:AB,AD,AF,AH) yield using XLSX data.
Skip non-numeric cells (like Google Sheets SUM does)."""
import requests, io, sys
import openpyxl
from decimal import Decimal

SHEET_ID = "1EO6-a_b9uun2vwETWu8aGh67ya3nwpdLAo4F-yjc1ZI"

print("Fetching XLSX...")
xlsx_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
resp = requests.get(xlsx_url, timeout=60)
wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
ws = wb['Evaluations']

# Find actual columns by letter
def col_num(letter):
    """Convert column letter to 1-based number."""
    return openpyxl.utils.column_index_from_string(letter)

# Show what's in columns I:M and T:Z at the header row
print("\n=== HEADER ROW (row 2) COLUMN CONTENTS ===")
for letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']:
    cn = col_num(letter)
    val = ws.cell(row=2, column=cn).value
    if val:
        print(f"  {letter:3s}: {val}")

# Header row 1 (merged header)
print("\n=== MERGED HEADER (row 1) ===")
for letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']:
    cn = col_num(letter)
    val = ws.cell(row=1, column=cn).value
    if val:
        print(f"  {letter:3s}: {val}")

# Now compute SUM(I:M) exactly as Google Sheets would
# SUM ignores text, booleans, and errors; only sums numbers
print(f"\n=== SUM(I:M) - Column by column ===")
im_total = Decimal('0')
for letter in ['I', 'J', 'K', 'L', 'M']:
    cn = col_num(letter)
    col_sum = Decimal('0')
    num_count = 0
    text_count = 0
    for row in range(3, ws.max_row + 1):  # Start after header row 2
        val = ws.cell(row=row, column=cn).value
        if val is not None:
            if isinstance(val, (int, float)):
                col_sum += Decimal(str(round(val, 2)))
                num_count += 1
            else:
                text_count += 1
    print(f"  SUM({letter}:{letter}) = {col_sum} ({num_count} numeric, {text_count} text)")
    im_total += col_sum
print(f"  SUM(I:M) TOTAL = {im_total}")

print(f"\n=== SUM(T:Z) - Column by column ===")
tz_total = Decimal('0')
for letter in ['T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
    cn = col_num(letter)
    col_sum = Decimal('0')
    num_count = 0
    text_count = 0
    for row in range(3, ws.max_row + 1):
        val = ws.cell(row=row, column=cn).value
        if val is not None:
            if isinstance(val, (int, float)):
                col_sum += Decimal(str(round(val, 2)))
                num_count += 1
            else:
                text_count += 1
    print(f"  SUM({letter}:{letter}) = {col_sum} ({num_count} numeric, {text_count} text)")
    tz_total += col_sum
print(f"  SUM(T:Z) TOTAL = {tz_total}")

print(f"\n=== B13 = SUM(I:M) + SUM(T:Z) = {im_total + tz_total} ===")
print(f"  Stats tab says: -26644.42")
print(f"  Our CSV calc:   -26646.11")

# Payouts
print(f"\n=== PAYOUTS: SUM(AC:AC) + SUM(AE:AE) + SUM(AG:AG) + SUM(AI:AI) ===")
# Sheet formula: =SUM(AB:AB)+SUM(AD:AD)+SUM(AF:AF)+SUM(AH:AH)  
# But XLSX has an extra column, so actual payout cols may be shifted!
# Let me check what columns AC, AE, AG, AI contain
for letter in ['AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']:
    cn = col_num(letter)
    header = ws.cell(row=2, column=cn).value
    print(f"  {letter}: {header}")

# Now compute payouts using the formula's column letters
print(f"\n=== SUM of payout columns (from formula AB,AD,AF,AH) ===")
payout_total = Decimal('0')
for letter in ['AB', 'AD', 'AF', 'AH']:
    cn = col_num(letter)
    col_sum = Decimal('0')
    num_count = 0
    for row in range(3, ws.max_row + 1):
        val = ws.cell(row=row, column=cn).value
        if val is not None and isinstance(val, (int, float)):
            col_sum += Decimal(str(round(val, 2)))
            num_count += 1
    header = ws.cell(row=2, column=cn).value
    print(f"  SUM({letter}:{letter}) [{header}] = {col_sum} ({num_count} values)")
    payout_total += col_sum
print(f"  PAYOUT TOTAL = {payout_total}")
print(f"  Stats tab says: 145295.20")
print(f"  Our CSV calc:   145295.62")

# Also check: Challenge Fees = -SUM(D:D)
cn_d = col_num('D')
fee_sum = Decimal('0')
for row in range(3, ws.max_row + 1):
    val = ws.cell(row=row, column=cn_d).value
    if val is not None and isinstance(val, (int, float)):
        fee_sum += Decimal(str(round(val, 2)))
print(f"\n=== B12 = -SUM(D:D) = {-fee_sum} ===")
print(f"  Stats tab says: -61234.37")
