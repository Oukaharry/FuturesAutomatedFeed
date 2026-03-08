"""Read XLSX Evaluations tab with openpyxl (data_only=True) to get actual formula results,
then compare cell-by-cell with CSV values."""
import requests, io, sys, json, sqlite3
import openpyxl
from decimal import Decimal

SHEET_ID = "1EO6-a_b9uun2vwETWu8aGh67ya3nwpdLAo4F-yjc1ZI"
sys.path.insert(0, '.')
from utils.data_processor import parse_currency

# Fetch XLSX
print("Fetching XLSX...")
xlsx_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
resp = requests.get(xlsx_url, timeout=60)
wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

# Get Evaluations sheet
ws = wb['Evaluations']
print(f"Evaluations sheet: {ws.max_row} rows x {ws.max_column} cols")

# Find header row
header_row = None
for row_idx in range(1, 15):
    for col_idx in range(1, 10):
        cell_val = ws.cell(row=row_idx, column=col_idx).value
        if cell_val and 'Prop Firm' in str(cell_val):
            header_row = row_idx
            break
    if header_row:
        break

print(f"Header at row: {header_row}")

# Read headers
headers = {}
for col_idx in range(1, ws.max_column + 1):
    val = ws.cell(row=header_row, column=col_idx).value
    if val:
        headers[col_idx] = str(val).strip()

# Print column mapping (for verification)
for col_idx in sorted(headers.keys())[:35]:
    letter = openpyxl.utils.get_column_letter(col_idx)
    print(f"  {letter} (col {col_idx}): {headers[col_idx]}")

# Identify hedge and payout columns by position
# Sheet formula B13: =SUM(I:M) + SUM(T:Z)
# I=9, M=13, T=20, Z=26 in 1-based column indexing
p1_hedge_range = list(range(9, 14))  # I through M
funded_hedge_range = list(range(20, 27))  # T through Z
# B15: =SUM(AB:AB)+SUM(AD:AD)+SUM(AF:AF)+SUM(AH:AH)
# AB=28, AD=30, AF=32, AH=34
payout_range = [28, 30, 32, 34]

# Read all data rows (after header)
data_start = header_row + 1

# Load stored evaluations for comparison
conn = sqlite3.connect('dashboard/dashboard.db')
conn.row_factory = sqlite3.Row
cur = conn.cursor()
cur.execute("SELECT evaluations FROM clients_data WHERE client_id='Tyler'")
data = cur.fetchone()
conn.close()
stored = json.loads(data['evaluations'])

# Sum XLSX values for columns I:M + T:Z (hedging)
xlsx_hedge = Decimal('0')
xlsx_payout = Decimal('0')
csv_hedge = Decimal('0')
csv_payout = Decimal('0')

row_count = 0
hedge_diffs = []
payout_diffs = []

for row_idx in range(data_start, ws.max_row + 1):
    # Check if row has a Prop Firm value
    prop_firm_col = None
    for ci, name in headers.items():
        if name == 'Prop Firm':
            prop_firm_col = ci
            break
    
    pf_val = ws.cell(row=row_idx, column=prop_firm_col).value if prop_firm_col else None
    if not pf_val or not str(pf_val).strip():
        continue
    
    # This is a valid data row
    if row_count >= len(stored):
        print(f"XLSX has more data rows than stored! ({row_count}+ vs {len(stored)})")
        break
    
    # Check hedge columns
    for ci in p1_hedge_range + funded_hedge_range:
        xlsx_val = ws.cell(row=row_idx, column=ci).value
        xlsx_num = round(float(xlsx_val), 2) if xlsx_val is not None and xlsx_val != '' else 0.0
        
        col_name = headers.get(ci, f'col{ci}')
        csv_raw = stored[row_count].get(col_name)
        csv_num = parse_currency(csv_raw)
        
        xlsx_hedge += Decimal(str(xlsx_num))
        csv_hedge += Decimal(str(csv_num))
        
        diff = Decimal(str(xlsx_num)) - Decimal(str(csv_num))
        if abs(diff) > Decimal('0.001'):
            hedge_diffs.append((row_count, col_name, xlsx_num, csv_num, float(diff)))
    
    # Check payout columns
    for ci in payout_range:
        xlsx_val = ws.cell(row=row_idx, column=ci).value
        xlsx_num = round(float(xlsx_val), 2) if xlsx_val is not None and xlsx_val != '' else 0.0
        
        col_name = headers.get(ci, f'col{ci}')
        csv_raw = stored[row_count].get(col_name)
        csv_num = parse_currency(csv_raw)
        
        xlsx_payout += Decimal(str(xlsx_num))
        csv_payout += Decimal(str(csv_num))
        
        diff = Decimal(str(xlsx_num)) - Decimal(str(csv_num))
        if abs(diff) > Decimal('0.001'):
            payout_diffs.append((row_count, col_name, xlsx_num, csv_num, float(diff)))
    
    row_count += 1

print(f"\nProcessed {row_count} data rows")

print(f"\n=== HEDGING ===")
print(f"XLSX sum:  {xlsx_hedge}")
print(f"CSV sum:   {csv_hedge}")
print(f"Diff:      {xlsx_hedge - csv_hedge}")
print(f"Stats tab: -26644.42")

if hedge_diffs:
    print(f"\nCell-level differences ({len(hedge_diffs)}):")
    for row, col, xval, cval, d in hedge_diffs:
        print(f"  Row {row}: {col}: xlsx={xval}, csv={cval}, diff={d:.4f}")
    total_diff = sum(d for _, _, _, _, d in hedge_diffs)
    print(f"Total cell diff: {total_diff:.4f}")

print(f"\n=== PAYOUTS ===")
print(f"XLSX sum:  {xlsx_payout}")
print(f"CSV sum:   {csv_payout}")
print(f"Diff:      {xlsx_payout - csv_payout}")
print(f"Stats tab: 145295.20")

if payout_diffs:
    print(f"\nCell-level differences ({len(payout_diffs)}):")
    for row, col, xval, cval, d in payout_diffs:
        print(f"  Row {row}: {col}: xlsx={xval}, csv={cval}, diff={d:.4f}")
    total_diff = sum(d for _, _, _, _, d in payout_diffs)
    print(f"Total cell diff: {total_diff:.4f}")
