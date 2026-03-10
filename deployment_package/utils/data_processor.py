from collections import defaultdict
from datetime import datetime
import pandas as pd
import requests
from io import StringIO
import math
import re

def clean_float(val):
    """
    Ensures float values are JSON compliant (no NaN or Inf).
    Returns 0.0 if invalid.
    """
    try:
        f_val = float(val)
        if math.isnan(f_val) or math.isinf(f_val):
            return 0.0
        return f_val
    except:
        return 0.0

def normalize_account_size(value):
    """
    Normalize account size values to standard format: $X,XXX (v1.0.1)
    
    Handles:
        - "50k", "50K" → "$50,000"
        - "100000", "100,000" → "$100,000"
        - "$50,000" → "$50,000" (already correct)
        - "5000" → "$5,000"
    """
    if not value:
        return value
    
    val = str(value).strip().upper()
    
    # Already in correct format
    if re.match(r'^\$[\d,]+$', val):
        return value
    
    # Handle "50k" or "50K" format
    match = re.match(r'^[\$]?(\d+\.?\d*)K$', val, re.IGNORECASE)
    if match:
        num = float(match.group(1)) * 1000
        return f"${num:,.0f}"
    
    # Handle plain numbers like "50000" or "50,000"
    val_clean = re.sub(r'[\$,\s]', '', val)
    try:
        num = float(val_clean)
        return f"${num:,.0f}"
    except:
        pass
    
    # Return original if can't parse
    return value

def clean_data_structure(data):
    """
    Recursively cleans a data structure (dict or list) to replace NaN/Inf with None.
    """
    if isinstance(data, dict):
        return {k: clean_data_structure(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [clean_data_structure(v) for v in data]
    elif isinstance(data, float):
        if math.isnan(data) or math.isinf(data):
            return None
        return data
    return data

def calculate_derived_metrics(df):
    """
    Applies Excel formulas to the DataFrame to ensure consistency with the Google Sheet.
    """
    # Helper to safely get float value
    def get_val(row, col):
        try:
            val = row.get(col)
            if pd.isna(val) or val == '':
                return 0.0
            # Handle string formatting if necessary
            if isinstance(val, str):
                s = val.replace('$', '').replace(' ', '').strip()
                if ',' in s:
                    if '.' not in s:
                        if re.search(r',\d{1,2}$', s):
                            return 0.0
                        s = s.replace(',', '')
                    else:
                        if re.search(r',\.', s):
                            return 0.0
                        s = s.replace(',', '')
                if s == '' or s == '-': return 0.0
                val = s
            return float(val)
        except:
            return 0.0

    # Helper to check if blank
    def is_blank(row, col):
        val = row.get(col)
        return pd.isna(val) or val == '' or str(val).strip() == ''

    for index, row in df.iterrows():
        # --- Column N: Hedge Net ---
        # Formula: =IF(OR(ISBLANK(I3), G3<>"Fail"), "", -D3 + I3 + J3+K3+L3+M3)
        # I=Hedge Result 1, G=Status P1, D=Fee
        
        status_p1 = str(row.get('Status P1', ''))
        
        if is_blank(row, 'Hedge Result 1') or status_p1 != 'Fail':
            df.at[index, 'Hedge Net'] = '' 
        else:
            fee = get_val(row, 'Fee')
            h1 = get_val(row, 'Hedge Result 1')
            h2 = get_val(row, 'Hedge Result 2')
            h3 = get_val(row, 'Hedge Result 3')
            h4 = get_val(row, 'Hedge Result 4')
            h5 = get_val(row, 'Hedge Result 5')
            
            # Formula: -Fee + Sum(Results)
            net = -fee + h1 + h2 + h3 + h4 + h5
            df.at[index, 'Hedge Net'] = net

        # --- Column AA: Hedge Net.1 ---
        # Formula:
        # IF S="Completed":
        #   SUM(Payouts) + SUM(Funded Hedge) + SUM(Phase 1 Hedge) - Fee - Activation Fee + SUM(Hedge Days)
        # IF S="Fail":
        #   SUM(Funded Hedge) + SUM(Phase 1 Hedge) - Fee - Activation Fee
        # ELSE: ""
        
        status = str(row.get('Status') or row.get('Status Funded', ''))
        
        # Sums
        sum_payouts = sum([get_val(row, c) for c in ['Payout 1', 'Payout 2', 'Payout 3', 'Payout 4']])
        
        sum_funded_hedge = sum([get_val(row, c) for c in [
            'Hedge Result 1.1', 'Hedge Result 2.1', 'Hedge Result 3.1', 
            'Hedge Result 4.1', 'Hedge Result 5.1', 'Hedge Result 6', 'Hedge Result 7'
        ]])
        
        sum_phase1_hedge = sum([get_val(row, c) for c in [
            'Hedge Result 1', 'Hedge Result 2', 'Hedge Result 3', 
            'Hedge Result 4', 'Hedge Result 5'
        ]])
        
        fee = get_val(row, 'Fee')
        activation_fee = get_val(row, 'Activation Fee')
        
        sum_hedge_days = sum([get_val(row, f'Hedge Day {i}') for i in range(1, 35)])  # All 34 farming days
        
        if status == 'Completed':
            # SUM(AB,AD,AF,AH, T,U,V,W,X,Y,Z, I,J,K,L,M) - D - P + SUM(AL...BN)
            val = sum_payouts + sum_funded_hedge + sum_phase1_hedge - fee - activation_fee + sum_hedge_days
            df.at[index, 'Hedge Net.1'] = val
            
        elif status == 'Fail':
            # T+U+V+W+X+Y+Z + I+J+K+L+M - D - P
            val = sum_funded_hedge + sum_phase1_hedge - fee - activation_fee
            df.at[index, 'Hedge Net.1'] = val
            
        else:
            df.at[index, 'Hedge Net.1'] = ''

    return df

def fetch_evaluations(sheet_url):
    """
    Fetches evaluation data from a public Google Sheet CSV export.
    Finds the header row dynamically by looking for 'Prop Firm'.
    """
    try:
        # Ensure we get CSV format
        if '/edit' in sheet_url:
            csv_url = sheet_url.replace('/edit?usp=sharing', '/export?format=csv').replace('/edit', '/export?format=csv')
        else:
            csv_url = sheet_url

        response = requests.get(csv_url)
        response.raise_for_status()
        
        # Read without header first to find the correct row
        df = pd.read_csv(StringIO(response.text), header=None)
        
        # Find the row that contains "Prop Firm" in the first few columns
        header_idx = -1
        for i, row in df.head(10).iterrows():
            # Check first 5 columns for "Prop Firm"
            if row.astype(str).str.contains('Prop Firm').any():
                header_idx = i
                break
        
        if header_idx != -1:
            # Reload with correct header
            df = pd.read_csv(StringIO(response.text), header=header_idx)
            
            # Clean up columns (remove unnamed, strip whitespace)
            df.columns = [str(c).strip() for c in df.columns]
            
            # Normalize column name variants: some sheets use 'Status Funded' instead of 'Status'
            if 'Status Funded' in df.columns and 'Status' not in df.columns:
                df = df.rename(columns={'Status Funded': 'Status'})
            
            # Define allowed columns based on the dashboard screenshot/template
            allowed_columns = [
                # Evaluation Info
                'Prop Firm', 'Account Size', 'Date Purchased', 'Fee',
                
                # Evaluation Phase
                'Date Started', 'Date Ended', 'Status P1', 'Account #',
                'Hedge Result 1', 'Hedge Result 2', 'Hedge Result 3', 'Hedge Result 4', 'Hedge Result 5', 'Hedge Net',
                
                # Funded Phase (duplicates get .1 suffix)
                'Account #.1', 'Activation Fee', 'Date Started.1', 'Date Ended.1', 'Status', 'Status Funded',
                'Hedge Result 1.1', 'Hedge Result 2.1', 'Hedge Result 3.1', 'Hedge Result 4.1', 'Hedge Result 5.1',
                'Hedge Result 6', 'Hedge Result 7', 'Hedge Net.1',
                'Payout 1', 'Date 1', 'Payout 2', 'Date 2', 'Payout 3', 'Date 3', 'Payout 4', 'Date 4',
                
                # Farming Phase
                'Farming Net'
            ]
            
            # Add Prop Day / Hedge Day 1-34 (sheet has 34 farming days)
            for i in range(1, 35):
                allowed_columns.append(f'Prop Day {i}')
                allowed_columns.append(f'Hedge Day {i}')
                
            # Filter to keep only allowed columns that exist in the dataframe
            existing_columns = [c for c in df.columns if c in allowed_columns]
            df = df[existing_columns]

            # Filter valid rows (where Prop Firm is not empty)
            if 'Prop Firm' in df.columns:
                df = df[df['Prop Firm'].notna()]
            
            # Apply derived metrics calculations
            df = calculate_derived_metrics(df)
            
            # Normalize Account Size values to standard format
            if 'Account Size' in df.columns:
                df['Account Size'] = df['Account Size'].apply(normalize_account_size)
                
            # Convert to list of dicts and clean recursively
            raw_data = df.to_dict(orient='records')
            return clean_data_structure(raw_data)
            
        return []
    except Exception as e:
        print(f"Error fetching sheet: {e}")
        return []


def fetch_stats_tab_values(sheet_url):
    """
    Fetches the Stats tab values from the Google Sheet XLSX export.
    Returns a dict with the cashflow and profitability values, or empty dict if unavailable.
    """
    try:
        import openpyxl, io, urllib.parse
        sheet_url = str(sheet_url).strip()
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', sheet_url)
        if not match:
            return {}
        sheet_key = match.group(1)
        xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=xlsx"
        resp = requests.get(xlsx_url, timeout=45)
        if resp.status_code != 200:
            return {}
        wb = openpyxl.load_workbook(filename=io.BytesIO(resp.content), data_only=True)
        if 'Stats' not in wb.sheetnames:
            return {}
        stats_ws = wb['Stats']
        stats_tab = {}
        current_section = None
        for row in stats_ws.iter_rows(min_row=1, max_row=30, max_col=2, values_only=True):
            label = str(row[0]).strip() if row[0] else ''
            val = row[1]
            if 'Profitability' in label and 'Completed' in label:
                current_section = 'profitability'
                continue
            elif 'Current Cashflow' in label:
                current_section = 'cashflow'
                continue
            elif 'Expected Value' in label:
                current_section = 'ev'
                continue
            elif 'Weekly' in label:
                current_section = None
                continue
            if current_section and label and val is not None and isinstance(val, (int, float)):
                key = label.lower().replace(' ', '_')
                if current_section == 'cashflow':
                    stats_tab[key] = round(float(val), 2)
                elif current_section == 'ev':
                    stats_tab[key] = round(float(val), 2)
                else:
                    stats_tab['prof_' + key] = round(float(val), 2)
            elif current_section and not label:
                current_section = None
        return {'__stats_tab__': stats_tab} if stats_tab else {}
    except Exception as e:
        print(f"Stats tab fetch failed (ignoring): {e}")
        return {}

def parse_currency(val):
    """
    Parses a currency string (e.g., '$100,000', '$1,234.56') to a float.
    Returns 0.0 for values Google Sheets cannot parse as numbers, so that
    our SUM logic matches sheet SUM() behaviour exactly.
    """
    if val is None:
        return 0.0
    try:
        s = str(val).replace('$', '').replace('\u20ac', '').replace('\u00a3', '').strip()
        if not s or s.lower() == 'nan':
            return 0.0
        if ',' in s:
            if '.' not in s:
                # No period: could be thousands ("1,234") or European decimal ("-573,79").
                # Google Sheets (US locale) cannot reliably parse these without a period,
                # so treat comma-decimal endings as unparseable (text → 0 in SUM).
                # Comma groups of exactly 3 digits are valid thousands separators.
                if re.search(r',\d{1,2}$', s):  # ends in ,X or ,XX  → ambiguous/euro → 0
                    return 0.0
                s = s.replace(',', '')           # thousands comma → strip safely
            else:
                # Has both comma and period.
                # Malformed if comma appears directly before period (e.g. "253,.96").
                if re.search(r',\.', s):
                    return 0.0
                s = s.replace(',', '')           # valid US thousands separator
        return round(float(s), 2)
    except:
        return 0.0


def calculate_statistics(evaluations, mt5_deals=None, mt5_account=None, xlsx_notes=None):
    """
    Calculates detailed statistics matching the 'Stats' sheet structure.
    Uses exact SUMIF logic from Google Sheet formulas.
    If xlsx_notes contains Stats tab values (from XLSX export), uses those
    for the cashflow section to guarantee exact match with Google Sheets.
    """
    stats = {
        "profitability_completed": {
            "challenge_fees": 0.0, "hedging_results": 0.0, "farming_results": 0.0, "payouts": 0.0, "net_profit": 0.0, "activation_fee": 0.0
        },
        "cashflow_inprogress": {
            "challenge_fees": 0.0, "hedging_results": 0.0, "farming_results": 0.0, "payouts": 0.0, "net_profit": 0.0, "activation_fee": 0.0
        },
        "evaluation_data": {},  # { "PropFirmName": { ...stats... } }
        "funded_data": {},      # { "PropFirmName": { ...stats... } }
        "eval_totals": {
            "total_running": 0, "not_started": 0, "ongoing": 0, "total_passed": 0, "total_failed": 0, "avg_net_failed": 0.0, "funded_rate": 0.0
        },
        "funded_totals": {
            "not_started": 0, "ongoing": 0, "failed": 0, "completed": 0, "avg_net_failed": 0.0, "avg_net_completed": 0.0, "total_funding": 0.0
        },
        "expected_value": 0.0,
        "ev_tracking": {
            "total_net_ended": 0.0,
            "count_ended": 0
        },
        "hedging_review": {
            "total_deposits": 0.0,
            "total_withdrawals": 0.0,
            "current_balance": 0.0,
            "actual_hedging_results": 0.0,
            "sheet_hedging_results": 0.0,
            "discrepancy": 0.0
        }
    }

    # If no evaluations and no MT5 data, return empty stats
    if not evaluations and not mt5_deals and not mt5_account:
        return stats
        
    evaluations = evaluations or []

    # Helper to get or create firm stats
    def get_firm_stats(firm_name, section):
        if firm_name not in stats[section]:
            stats[section][firm_name] = {
                "total_running": 0, "not_started": 0, "ongoing": 0, "total_passed": 0, "total_failed": 0, "net_failed_sum": 0.0,
                "failed": 0, "completed": 0, "net_completed_sum": 0.0, "total_funding": 0.0
            }
        return stats[section][firm_name]

    # Pre-populate target firms to match sheet structure
    TARGET_FIRMS = ["My Funded Futures", "Funding Ticks", "Trade Day", "Funded Next", "Top Step"]
    for firm in TARGET_FIRMS:
        get_firm_stats(firm, "evaluation_data")
        get_firm_stats(firm, "funded_data")

    sheet_hedge_total = 0.0
    
    # Phase 1 hedge columns (J-N in sheet = Hedge Result 1-5)
    P1_HEDGE_COLS = ['Hedge Result 1', 'Hedge Result 2', 'Hedge Result 3', 'Hedge Result 4', 'Hedge Result 5']
    # Funded hedge columns (U-AA in sheet = Hedge Result 1.1-7)
    FUNDED_HEDGE_COLS = ['Hedge Result 1.1', 'Hedge Result 2.1', 'Hedge Result 3.1', 'Hedge Result 4.1', 
                         'Hedge Result 5.1', 'Hedge Result 6', 'Hedge Result 7']
    # Hedge Day columns for farming
    HEDGE_DAY_COLS = [f'Hedge Day {i}' for i in range(1, 35)]

    for ev in evaluations:
        firm = ev.get('Prop Firm', 'Unknown')
        status_p1 = str(ev.get('Status P1', '')).strip()
        # Support both 'Status' and 'Status Funded' column names (sheet variant)
        status_funded = str(ev.get('Status') or ev.get('Status Funded', '')).strip()
        
        # Normalize for comparison (but keep original for exact match)
        status_p1_lower = status_p1.lower()
        status_funded_lower = status_funded.lower()
        
        # === CALCULATE VALUES USING EXACT SHEET SUMIF LOGIC ===
        
        # Get individual hedge results (NOT the calculated Hedge Net columns)
        # Round each row-level sum to 2dp to match Google Sheets cell precision
        p1_hedges = round(sum(parse_currency(ev.get(col)) for col in P1_HEDGE_COLS), 2)
        funded_hedges = round(sum(parse_currency(ev.get(col)) for col in FUNDED_HEDGE_COLS), 2)
        hedge_days = round(sum(parse_currency(ev.get(col)) for col in HEDGE_DAY_COLS), 2)
        
        fee = parse_currency(ev.get('Fee'))
        activation_fee = parse_currency(ev.get('Activation Fee'))
        farming_net = parse_currency(ev.get('Farming Net'))
        payouts = round(sum(parse_currency(ev.get(f'Payout {i}')) for i in range(1, 5)), 2)
        
        # For sheet_hedge_total tracking
        hedge_net = parse_currency(ev.get('Hedge Net')) + parse_currency(ev.get('Hedge Net.1'))
        p1_hedge_net = parse_currency(ev.get('Hedge Net'))
        funded_hedge_net = parse_currency(ev.get('Hedge Net.1'))
        sheet_hedge_total += hedge_net
        
        # EV tracking: Sheet formula = SUM(O:O, AB:AB) / COUNT(O:O, AB:AB)
        # O = Hedge Net (P1), AB = Hedge Net.1 (Funded)
        # Count each non-empty cell independently (a row can contribute to both)
        raw_p1_hn = ev.get('Hedge Net')
        raw_fd_hn = ev.get('Hedge Net.1')
        if raw_p1_hn is not None and str(raw_p1_hn).strip() not in ('', '-'):
            stats["ev_tracking"]["total_net_ended"] += p1_hedge_net
            stats["ev_tracking"]["count_ended"] += 1
        if raw_fd_hn is not None and str(raw_fd_hn).strip() not in ('', '-'):
            stats["ev_tracking"]["total_net_ended"] += funded_hedge_net
            stats["ev_tracking"]["count_ended"] += 1
        
        # === CASHFLOW - IN PROGRESS (TOTALS of ALL records - no status filtering) ===
        # Sheet formulas SUM entire columns, including deleted rows:
        # Challenge Fees: =-SUM(Evaluations!D:D) = negative sum of ALL Fee column only
        # Hedging Results: =SUM(J:N) + SUM(U:AA) = ALL hedge results
        # Farming Results: =SUM(all Hedge Day columns) = ALL farming
        # Payouts: =SUM(AC,AE,AG,AI) = ALL payouts
        stats["cashflow_inprogress"]["challenge_fees"] = round(stats["cashflow_inprogress"]["challenge_fees"] + fee + activation_fee, 2)
        stats["cashflow_inprogress"]["hedging_results"] = round(stats["cashflow_inprogress"]["hedging_results"] + p1_hedges + funded_hedges, 2)
        stats["cashflow_inprogress"]["farming_results"] = round(stats["cashflow_inprogress"]["farming_results"] + hedge_days, 2)
        stats["cashflow_inprogress"]["payouts"] = round(stats["cashflow_inprogress"]["payouts"] + payouts, 2)
        stats["cashflow_inprogress"]["activation_fee"] = round(stats["cashflow_inprogress"]["activation_fee"] + activation_fee, 2)
        
        # Skip deleted accounts for profitability and other sections
        if 'deleted' in status_p1_lower or 'deleted' in status_funded_lower:
            continue
        
        # === PROFITABILITY - COMPLETED (exact SUMIF logic from sheet) ===
        # Challenge Fees formula: (SUMIF(Fee,P1="Fail") + SUMIF(Fee,Status="Completed") + SUMIF(Fee,Status="Fail")) * -1
        # We store as positive, the display will show as negative
        
        # Hedging Results Completed formula:
        # Part 1: P1 hedges (J-N) where Status P1 = "Fail"
        # Part 2: Funded hedges (U-AA) where Status = "Fail" or "Completed"  
        # Part 3: P1 hedges (J-N) where Status = "Fail" or "Completed"
        
        # Farming Results Completed formula:
        # Sum of Hedge Day columns (AM,AO,AQ...) where Status = "Completed" ONLY
        
        is_p1_fail = status_p1 == 'Fail'
        is_funded_fail = status_funded == 'Fail'
        is_funded_completed = status_funded == 'Completed'
        is_funded_ended = is_funded_fail or is_funded_completed
        
        # Challenge Fees Completed: Fee + Activation Fee where P1=Fail OR Status=Fail OR Status=Completed
        if is_p1_fail:
            stats["profitability_completed"]["challenge_fees"] = round(stats["profitability_completed"]["challenge_fees"] + fee + activation_fee, 2)
        if is_funded_fail:
            stats["profitability_completed"]["challenge_fees"] = round(stats["profitability_completed"]["challenge_fees"] + fee + activation_fee, 2)
        if is_funded_completed:
            stats["profitability_completed"]["challenge_fees"] = round(stats["profitability_completed"]["challenge_fees"] + fee + activation_fee, 2)
            
        # Hedging Results Completed
        if is_p1_fail:
            stats["profitability_completed"]["hedging_results"] = round(stats["profitability_completed"]["hedging_results"] + p1_hedges, 2)
        if is_funded_ended:
            stats["profitability_completed"]["hedging_results"] = round(stats["profitability_completed"]["hedging_results"] + funded_hedges + p1_hedges, 2)
            
        # Farming Results Completed: Hedge Days ONLY where Status=Completed
        if is_funded_completed:
            stats["profitability_completed"]["farming_results"] = round(stats["profitability_completed"]["farming_results"] + hedge_days, 2)
            
        # Payouts Completed: where Status=Completed ONLY (sheet formula: SUMIF Status="Completed")
        if is_funded_completed:
            stats["profitability_completed"]["payouts"] = round(stats["profitability_completed"]["payouts"] + payouts, 2)
            
        # Activation Fee for Completed (B25 in Net Profit formula)
        # Track activation fee for accounts that have ended
        if is_p1_fail or is_funded_ended:
            stats["profitability_completed"]["activation_fee"] = round(stats["profitability_completed"]["activation_fee"] + activation_fee, 2)
        
        # Track "in progress" for evaluation data section (accounts not ended)
        is_in_progress = not is_p1_fail and not is_funded_ended

        # --- 2. Evaluation Data ---
        estats = get_firm_stats(firm, "evaluation_data")
        
        # Calculate net_profit for this row
        total_fee = fee + activation_fee
        total_hedge = p1_hedges + funded_hedges
        total_farm = farming_net + hedge_days
        net_profit = payouts + total_hedge + total_farm - total_fee
        
        # Eval Status Logic
        # Active/Ongoing
        if is_in_progress:
            estats["total_running"] += 1
            stats["eval_totals"]["total_running"] += 1
            
            if not ev.get('Date Started'):
                estats["not_started"] += 1
                stats["eval_totals"]["not_started"] += 1
            else:
                estats["ongoing"] += 1
                stats["eval_totals"]["ongoing"] += 1
                
        # Passed (P1)
        if status_p1 == 'Pass' or status_p1_lower == 'pass':
            estats["total_passed"] += 1
            stats["eval_totals"]["total_passed"] += 1
            
        # Failed (P1)
        if is_p1_fail:
            estats["total_failed"] += 1
            stats["eval_totals"]["total_failed"] += 1
            estats["net_failed_sum"] += net_profit 
            stats["eval_totals"]["avg_net_failed"] += net_profit
            
            # EV tracking moved to cashflow section above (unconditional, matches Sheet SUM/COUNT)

        # --- 3. Funded Data ---
        fstats = get_firm_stats(firm, "funded_data")
        
        # Only count if it reached funded stage (Status is present or P1 passed)
        has_funded_status = status_funded and status_funded != '-'
        p1_passed = status_p1 == 'Pass' or status_p1_lower == 'pass'
        
        if has_funded_status or p1_passed:
            funding_amount = parse_currency(ev.get('Account Size'))
            
            # Total Funding: Sum of ALL accounts that reached funded stage
            fstats["total_funding"] += funding_amount
            stats["funded_totals"]["total_funding"] += funding_amount

            if is_in_progress:
                if not ev.get('Date Started.1'): # Funded Start Date
                     fstats["not_started"] += 1 
                     stats["funded_totals"]["not_started"] += 1
                else:
                    fstats["ongoing"] += 1
                    stats["funded_totals"]["ongoing"] += 1

            elif is_funded_fail:
                fstats["failed"] += 1
                stats["funded_totals"]["failed"] += 1
                fstats["net_failed_sum"] += net_profit
                stats["funded_totals"]["avg_net_failed"] += net_profit
                
                # EV tracking moved to cashflow section above (unconditional, matches Sheet SUM/COUNT)
                
            elif is_funded_completed:
                fstats["completed"] += 1
                stats["funded_totals"]["completed"] += 1
                fstats["net_completed_sum"] += net_profit
                stats["funded_totals"]["avg_net_completed"] += net_profit

    # --- Calculate Averages & Rates ---
    
    # Eval Averages
    if stats["eval_totals"]["total_failed"] > 0:
        stats["eval_totals"]["avg_net_failed"] /= stats["eval_totals"]["total_failed"]
    
    total_eval_closed = stats["eval_totals"]["total_passed"] + stats["eval_totals"]["total_failed"]
    if total_eval_closed > 0:
        stats["eval_totals"]["funded_rate"] = (stats["eval_totals"]["total_passed"] / total_eval_closed) * 100

    # Funded Averages
    if stats["funded_totals"]["failed"] > 0:
        stats["funded_totals"]["avg_net_failed"] /= stats["funded_totals"]["failed"]
        
    if stats["funded_totals"]["completed"] > 0:
        stats["funded_totals"]["avg_net_completed"] /= stats["funded_totals"]["completed"]

    # Per-Firm Averages
    for firm in stats["evaluation_data"]:
        d = stats["evaluation_data"][firm]
        if d["total_failed"] > 0:
            d["avg_net_failed"] = d["net_failed_sum"] / d["total_failed"]
        else:
            d["avg_net_failed"] = 0.0
            
        closed = d["total_passed"] + d["total_failed"]
        d["funded_rate"] = (d["total_passed"] / closed * 100) if closed > 0 else 0.0

    for firm in stats["funded_data"]:
        d = stats["funded_data"][firm]
        if d["failed"] > 0:
            d["avg_net_failed"] = d["net_failed_sum"] / d["failed"]
        else:
            d["avg_net_failed"] = 0.0
            
        if d["completed"] > 0:
            d["avg_net_completed"] = d["net_completed_sum"] / d["completed"]
        else:
            d["avg_net_completed"] = 0.0

    # --- Expected Value ---
    # Sheet formula: =IFERROR(SUM(Evaluations!O:O,Evaluations!AB:AB) / COUNT(Evaluations!O:O,Evaluations!AB:AB),0)
    # O = Hedge Net (P1), AB = Hedge Net.1 (Funded) — average of non-empty hedge net cells
    ev_data = stats["ev_tracking"]
    if ev_data["count_ended"] > 0:
        stats["expected_value"] = round(ev_data["total_net_ended"] / ev_data["count_ended"], 2)
    else:
        stats["expected_value"] = 0.0

    # --- Hedging Review ---
    # Sheet Hedging Results = Total Hedging Results + Total Farming Results (from ALL evaluations)
    # cashflow_inprogress already contains the sum of ALL records (no status filter)
    # So we use those totals directly (not adding completed which would double-count)
    total_hedging = stats["cashflow_inprogress"]["hedging_results"]
    total_farming = stats["cashflow_inprogress"]["farming_results"]
    stats["hedging_review"]["sheet_hedging_results"] = total_hedging + total_farming
    
    # Debug logging (will be visible in server logs)
    import sys
    debug_log = []
    
    if mt5_account:
        # Handle both dict (serialized) and object
        if isinstance(mt5_account, dict):
            balance = float(mt5_account.get('balance', 0.0) or 0.0)
            deposits = float(mt5_account.get('total_deposits', 0.0) or 0.0)
            withdrawals = float(mt5_account.get('total_withdrawals', 0.0) or 0.0)
        else:
            balance = float(getattr(mt5_account, 'balance', 0.0) or 0.0)
            deposits = float(getattr(mt5_account, 'total_deposits', 0.0) or 0.0)
            withdrawals = float(getattr(mt5_account, 'total_withdrawals', 0.0) or 0.0)
        
        stats["hedging_review"]["current_balance"] = balance
        stats["hedging_review"]["total_deposits"] = deposits
        stats["hedging_review"]["total_withdrawals"] = withdrawals
        
        # Calculate Actual Hedging Results using the Google Sheet formula:
        # =IF(AND(B20<>"", B22<>""), B22-(B20-B21), "")
        # Which is: Current Balance - (Total Deposits - Total Withdrawals)
        # Note: withdrawals are already negative, so we add them
        net_deposits = deposits + withdrawals  # withdrawals is negative
        actual_hedging = balance - net_deposits
        stats["hedging_review"]["actual_hedging_results"] = actual_hedging
        stats["hedging_review"]["discrepancy"] = actual_hedging - stats["hedging_review"]["sheet_hedging_results"]
        
        debug_log.append(f"MT5 Account: balance=${balance:.2f}, deposits=${deposits:.2f}, withdrawals=${withdrawals:.2f}")
        debug_log.append(f"Calculated: net_deposits=${net_deposits:.2f}, actual_hedging=${actual_hedging:.2f}")
        has_mt5_data = True
    else:
        debug_log.append("MT5 Account: NONE")
        has_mt5_data = False

    # Process deals if available (for trade history tracking only, not hedging review)
    if mt5_deals and len(mt5_deals) > 0:
        deal_types_seen = set()
        balance_count = 0
        trade_count = 0
        actual_profit = 0.0
        
        debug_log.append(f"MT5 Deals: {len(mt5_deals)} total")
        
        for deal in mt5_deals:
            # Handle both dict (serialized) and object
            if isinstance(deal, dict):
                d_type = deal.get('type')
                d_profit = float(deal.get('profit', 0.0) or 0.0)
                d_swap = float(deal.get('swap', 0.0) or 0.0)
                d_comm = float(deal.get('commission', 0.0) or 0.0)
            else:
                d_type = getattr(deal, 'type', None)
                d_profit = float(getattr(deal, 'profit', 0.0) or 0.0)
                d_swap = float(getattr(deal, 'swap', 0.0) or 0.0)
                d_comm = float(getattr(deal, 'commission', 0.0) or 0.0)
            
            deal_types_seen.add(str(d_type))
            
            # DEAL_TYPE_BALANCE = 2 or "BALANCE" (serialized from trader app)
            is_balance = d_type == 2 or str(d_type).upper() == "BALANCE"
            if is_balance:
                balance_count += 1
            else:
                trade_count += 1
                actual_profit += (d_profit + d_swap + d_comm)
        
        if not has_mt5_data:
            has_mt5_data = True
        
        # Debug: store deal types seen for troubleshooting
        debug_log.append(f"   - Balance deals: {balance_count}, Trade deals: {trade_count}")
        debug_log.append(f"   - Trade profit from deals: ${actual_profit:.2f}")
        debug_log.append(f"   - Deal types seen: {list(deal_types_seen)}")
        
        stats["hedging_review"]["_debug_deal_count"] = len(mt5_deals)
        stats["hedging_review"]["_debug_deal_types"] = list(deal_types_seen)
        stats["hedging_review"]["_debug_balance_count"] = balance_count
    else:
        debug_log.append("MT5 Deals: NONE or empty")
    
    # Print debug log to console/logs
    print("\n🔍 DATA_PROCESSOR DEBUG:")
    for line in debug_log:
        print(f"   {line}")
    print()

    # --- Override cashflow with Stats tab values if available ---
    stats_tab = (xlsx_notes or {}).get('__stats_tab__', {})
    if stats_tab:
        cf = stats["cashflow_inprogress"]
        if 'challenge_fees' in stats_tab:
            cf['challenge_fees'] = abs(stats_tab['challenge_fees'])
        if 'hedging_results' in stats_tab:
            cf['hedging_results'] = stats_tab['hedging_results']
        if 'farming_results' in stats_tab:
            cf['farming_results'] = stats_tab['farming_results']
        if 'payouts' in stats_tab:
            cf['payouts'] = stats_tab['payouts']

        pc = stats["profitability_completed"]
        if 'prof_challenge_fees' in stats_tab:
            pc['challenge_fees'] = abs(stats_tab['prof_challenge_fees'])
        if 'prof_hedging_results' in stats_tab:
            pc['hedging_results'] = stats_tab['prof_hedging_results']
        if 'prof_farming_results' in stats_tab:
            pc['farming_results'] = stats_tab['prof_farming_results']
        if 'prof_payouts' in stats_tab:
            pc['payouts'] = stats_tab['prof_payouts']

        # Override EV with Stats tab value if available
        if 'ev' in stats_tab:
            stats['expected_value'] = stats_tab['ev']

    # --- Calculate Net Profit for each section (AFTER discrepancy is calculated) ---
    # Formula: Net Profit = Payouts + Challenge Fees (neg) + Hedging + Farming + Discrepancy
    # The sheet formula is: =B6+B3+B4+B5+B25 
    # where B3=fees(negative), B4=hedging, B5=farming, B6=payouts, B25=discrepancy
    # Discrepancy = Actual Hedging Results - Sheet Hedging Results (from MT5)
    # NOTE: Only include discrepancy if we have actual MT5 data, otherwise use 0
    discrepancy = stats["hedging_review"]["discrepancy"] if has_mt5_data else 0.0
    
    for section in ["profitability_completed", "cashflow_inprogress"]:
        s = stats[section]
        # fees are stored positive but are expenses (subtract them)
        # Add discrepancy (B25) which adjusts for MT5 actual vs sheet recorded
        s["net_profit"] = s["payouts"] + s["hedging_results"] + s["farming_results"] - s["challenge_fees"] + discrepancy

    # Rounding
    def round_dict(d):
        for k, v in d.items():
            if isinstance(v, float):
                d[k] = round(v, 2)
            elif isinstance(v, dict):
                round_dict(v)
                
    round_dict(stats)
    return stats

def extract_unique_values(data):
    """
    Extracts unique values for dropdown fields from the data.
    Merges with default options to ensure a good baseline.
    """
    # Default options (baseline)
    options = {
        'Prop Firm': {'MFFU', 'MFFU_Flex', 'Funded Next', 'Funding Ticks', 'Top Step', 'Lucid', 'Trade Day', 'Alpha Futures', 'Tradeify', 'Other'},
        'Account Size': {'$5,000', '$10,000', '$25,000', '$50,000', '$100,000', '$200,000'},
        'Status': {'Active', 'Passed', 'Breached', 'Closed', 'Payout'}
    }
    
    if not data:
        return {k: sorted(list(v)) for k, v in options.items()}

    for row in data:
        # Prop Firm
        val = row.get('Prop Firm')
        if val: options['Prop Firm'].add(str(val).strip())
            
        # Account Size - normalize to standard format
        val = row.get('Account Size')
        if val: 
            normalized = normalize_account_size(val)
            options['Account Size'].add(normalized)
            
        # Status (P1)
        val = row.get('Status P1')
        if val: options['Status'].add(str(val).strip())
            
        # Status (Funded)
        val = row.get('Status')
        if val: options['Status'].add(str(val).strip())
    
    # Return sorted lists
    return {k: sorted(list(v)) for k, v in options.items()}

def group_deals_by_position(deals):
    """
    Groups a list of MT5 deals by position_id to form complete trades.
    """
    positions = defaultdict(list)
    
    for deal in deals:
        # Access attributes safely (handle both object and dict)
        try:
            if isinstance(deal, dict):
                pos_id = deal.get('position_id')
            else:
                pos_id = getattr(deal, 'position_id', None)
        except:
            continue
            
        if pos_id is not None:
            positions[pos_id].append(deal)
            
    trades = []
    for pos_id, pos_deals in positions.items():
        if not pos_deals:
            continue
            
        # Sort by time to find entry and exit
        # Handle both object and dict access for 'time'
        pos_deals.sort(key=lambda x: x.get('time') if isinstance(x, dict) else getattr(x, 'time', 0))
        
        first_deal = pos_deals[0]
        last_deal = pos_deals[-1]
        
        # Helper to get value
        def get_val(obj, attr, default=0.0):
            if isinstance(obj, dict):
                return obj.get(attr, default)
            return getattr(obj, attr, default)

        # Aggregate values
        total_profit = sum(get_val(d, 'profit') for d in pos_deals)
        total_swap = sum(get_val(d, 'swap') for d in pos_deals)
        total_commission = sum(get_val(d, 'commission') for d in pos_deals)
        net_profit = total_profit + total_swap + total_commission
        
        # Basic info from first deal (Entry)
        symbol = get_val(first_deal, 'symbol', '')
        deal_type = get_val(first_deal, 'type', 0) 
        volume = get_val(first_deal, 'volume', 0.0)
        open_time = get_val(first_deal, 'time', 0)
        open_price = get_val(first_deal, 'price', 0.0)
        
        # Exit info
        close_time = get_val(last_deal, 'time', 0)
        close_price = get_val(last_deal, 'price', 0.0)
        
        trades.append({
            "position_id": pos_id,
            "symbol": symbol,
            "type": "BUY" if deal_type == 0 else "SELL",
            "volume": clean_float(volume),
            "open_time": datetime.fromtimestamp(open_time).strftime('%Y-%m-%d %H:%M'),
            "open_price": clean_float(open_price),
            "close_time": datetime.fromtimestamp(close_time).strftime('%Y-%m-%d %H:%M'),
            "close_price": clean_float(close_price),
            "net_profit": clean_float(round(net_profit, 2))
        })
        
    return trades

def serialize_positions(positions):
    """
    Converts MT5 position objects to dictionaries.
    """
    data = []
    if positions is None:
        return data
        
    for pos in positions:
        try:
            data.append({
                "ticket": getattr(pos, 'ticket', 0),
                "symbol": getattr(pos, 'symbol', ''),
                "type": "BUY" if getattr(pos, 'type', 0) == 0 else "SELL",
                "volume": clean_float(getattr(pos, 'volume', 0.0)),
                "open_price": clean_float(getattr(pos, 'price_open', 0.0)),
                "current_price": clean_float(getattr(pos, 'price_current', 0.0)),
                "sl": clean_float(getattr(pos, 'sl', 0.0)),
                "tp": clean_float(getattr(pos, 'tp', 0.0)),
                "swap": clean_float(getattr(pos, 'swap', 0.0)),
                "profit": clean_float(getattr(pos, 'profit', 0.0))
            })
        except:
            continue
    return data

def serialize_account_info(account):
    """
    Converts MT5 account info object to dictionary.
    """
    if account is None:
        return {}
        
    try:
        return {
            "login": getattr(account, 'login', 0),
            "balance": clean_float(getattr(account, 'balance', 0.0)),
            "equity": clean_float(getattr(account, 'equity', 0.0)),
            "profit": clean_float(getattr(account, 'profit', 0.0)),
            "margin": clean_float(getattr(account, 'margin', 0.0)),
            "margin_free": clean_float(getattr(account, 'margin_free', 0.0)),
            "margin_level": clean_float(getattr(account, 'margin_level', 0.0))
        }
    except:
        return {}
