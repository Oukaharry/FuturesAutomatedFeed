import requests
import openpyxl
import os

# URL to export as XLSX
sheet_id = "1NX46wyWWGVOyb9IyTAEnjQUKfQ6A53Yr8MazhIJVOAY"
url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
filename = "sheet_dump.xlsx"

print(f"Downloading {url}...")
response = requests.get(url)

if response.status_code == 200:
    with open(filename, 'wb') as f:
        f.write(response.content)
    print(f"Saved to {filename}")
    
    try:
        print("Loading workbook...")
        wb = openpyxl.load_workbook(filename, data_only=False) # data_only=False loads formulas
        
        all_formulas = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\nProcessing sheet: {sheet_name}")
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_info = {
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": cell.value
                        }
                        all_formulas.append(formula_info)
        
        with open('extracted_formulas.txt', 'w', encoding='utf-8') as f:
            for item in all_formulas:
                f.write(f"Sheet: {item['sheet']}, Cell: {item['cell']}, Formula: {item['formula']}\n")
        
        print(f"\nFound {len(all_formulas)} formulas. Saved to extracted_formulas.txt")
        
    except Exception as e:
        print(f"Error processing workbook: {e}")

else:
    print(f"Failed to download sheet. Status code: {response.status_code}")
