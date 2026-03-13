"""
Re-import Google Sheet cell comments into cell_notes for all clients.

This extracts cell comments (notes) from each client's Google Sheet XLSX export
and saves them as cell_notes, which are displayed in the Prop Progress columns.

Usage:
    python reimport_sheet_notes.py           # Dry run - show what would be imported
    python reimport_sheet_notes.py --apply   # Actually import notes
    python reimport_sheet_notes.py --client Ed  # Single client only
"""
import sys
import os
import json
import sqlite3
import re
import requests
import io

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

DB_PATH = os.path.join('dashboard', 'dashboard.db')


def get_all_clients_with_sheets():
    """Get all clients that have a sheet_url in their identity."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT client_id, identity FROM clients_data WHERE identity IS NOT NULL")
    clients = []
    for row in cur.fetchall():
        client_id = row[0]
        try:
            identity = json.loads(row[1]) if isinstance(row[1], str) else row[1]
            sheet_url = identity.get('sheet_url', '') if isinstance(identity, dict) else ''
            if sheet_url and 'docs.google.com/spreadsheets' in sheet_url:
                clients.append((client_id, sheet_url))
        except (json.JSONDecodeError, AttributeError):
            pass
    conn.close()
    return clients


def fetch_xlsx_comments(sheet_url):
    """Extract cell comments from Google Sheet XLSX export."""
    import openpyxl
    
    notes = {}
    sheet_url = str(sheet_url).strip()
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', sheet_url)
    if not match:
        return notes
    
    sheet_key = match.group(1)
    xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=xlsx&gid=0"
    
    resp = requests.get(xlsx_url, timeout=90)
    if resp.status_code != 200:
        print(f"    XLSX fetch failed: HTTP {resp.status_code}")
        return notes
    
    wb = openpyxl.load_workbook(filename=io.BytesIO(resp.content), data_only=True)
    ws = wb.active
    
    header_idx = -1
    col_map = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False)):
        row_vals = [str(c.value).strip() if c.value else '' for c in row]
        if any('Prop Firm' in str(v) for v in row_vals):
            header_idx = r_idx
            col_map = {idx: str(h).strip() for idx, h in enumerate(row_vals) if h}
            break
        elif any('Account Size' in str(v) for v in row_vals):
            header_idx = r_idx
            col_map = {idx: str(h).strip() for idx, h in enumerate(row_vals) if h}
            if 0 not in col_map or not col_map[0]:
                col_map[0] = 'Prop Firm'
            break
    
    if header_idx == -1:
        print("    Could not find header row")
        return notes
    
    data_row_counter = 0
    for row_cells in ws.iter_rows(min_row=header_idx + 2, values_only=False):
        is_valid = False
        for c_idx, cell in enumerate(row_cells):
            if c_idx in col_map and col_map[c_idx] == 'Prop Firm':
                if cell.value and str(cell.value).strip():
                    is_valid = True
                break
        if is_valid:
            for c_idx, cell in enumerate(row_cells):
                if c_idx in col_map and cell.comment:
                    c_name = col_map[c_idx]
                    if data_row_counter not in notes:
                        notes[data_row_counter] = {}
                    notes[data_row_counter][c_name] = cell.comment.text.strip()
            data_row_counter += 1
    
    return notes


def save_notes_to_db(client_id, notes, dry_run=True, quiet=False):
    """Save extracted notes to cell_notes table."""
    if not notes:
        return 0
    
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    # Ensure table exists
    cur.execute('''
        CREATE TABLE IF NOT EXISTS cell_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            column_key TEXT NOT NULL,
            note_content TEXT NOT NULL,
            created_by TEXT DEFAULT 'sheet_import',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(client_id, row_index, column_key)
        )
    ''')
    
    saved = 0
    skipped = 0
    for row_idx, col_notes in notes.items():
        if not isinstance(row_idx, int) or not isinstance(col_notes, dict):
            continue
        for col_key, content in col_notes.items():
            if not content or not str(content).strip():
                continue
            content = str(content).strip()
            
            if dry_run:
                if not quiet:
                    print(f"    [DRY] row {row_idx:>4} | {col_key:<20} | {content[:50]}")
                saved += 1
            else:
                try:
                    cur.execute('''
                        INSERT OR REPLACE INTO cell_notes 
                        (client_id, row_index, column_key, note_content, created_by, updated_at)
                        VALUES (?, ?, ?, ?, 'sheet_import', CURRENT_TIMESTAMP)
                    ''', (client_id, row_idx, col_key, content))
                    saved += 1
                except Exception as e:
                    print(f"    ERROR saving row {row_idx} {col_key}: {e}")
                    skipped += 1
    
    if not dry_run:
        conn.commit()
    conn.close()
    return saved


def main():
    dry_run = '--apply' not in sys.argv
    summary_only = '--summary' in sys.argv
    single_client = None
    
    if '--client' in sys.argv:
        idx = sys.argv.index('--client')
        if idx + 1 < len(sys.argv):
            single_client = sys.argv[idx + 1]
    
    if dry_run:
        print("=== DRY RUN (use --apply to save) ===\n")
    else:
        print("=== APPLYING CHANGES ===\n")
    
    clients = get_all_clients_with_sheets()
    if single_client:
        clients = [(cid, url) for cid, url in clients if cid == single_client]
    
    print(f"Found {len(clients)} clients with sheet URLs\n")
    
    total_notes = 0
    for client_id, sheet_url in clients:
        print(f"--- {client_id} ---")
        print(f"  Sheet: {sheet_url[:80]}...")
        
        try:
            notes = fetch_xlsx_comments(sheet_url)
            note_count = sum(len(v) for v in notes.values() if isinstance(v, dict))
            print(f"  Found {note_count} cell comments across {len(notes)} rows")
            
            if notes:
                saved = save_notes_to_db(client_id, notes, dry_run=dry_run, quiet=summary_only)
                total_notes += saved
                if not dry_run:
                    print(f"  Saved {saved} notes")
        except Exception as e:
            print(f"  ERROR: {e}")
        print()
    
    print(f"\n{'Would save' if dry_run else 'Saved'} {total_notes} total notes")
    if dry_run:
        print("\nRun with --apply to actually save the notes.")


if __name__ == '__main__':
    main()
