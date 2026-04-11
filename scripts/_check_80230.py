"""Quick check: does account 80230 appear anywhere in the Google Sheet?"""
import urllib.request, io, sys
try:
    import openpyxl
except ImportError:
    print("Need openpyxl"); sys.exit(1)

sheet_key = '1q4atojmjW03XLU6bRfubZ3WZiK071x3eQttt5kdKVYs'
url = f'https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=xlsx'
req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
resp = urllib.request.urlopen(req, timeout=60)
wb = openpyxl.load_workbook(filename=io.BytesIO(resp.read()), data_only=True)

# Search all sheets for "80230"
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for c_idx, cell in enumerate(row):
            if cell and '80230' in str(cell):
                print(f"FOUND '80230' in sheet '{sheet_name}' row {r_idx} col {c_idx+1}: {cell}")

# Also check V2-2641 in detail - the extra day issue
ws = wb[wb.sheetnames[0]]
header_idx = None
col_map = {}
for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
    row_vals = [str(c).strip() if c else '' for c in row]
    if any('Prop Firm' in v for v in row_vals):
        header_idx = r_idx
        for ci, h in enumerate(row_vals):
            if h: col_map[ci] = h
        break

# Find the row for V2-2641/6337 and print ALL its values especially account columns
if header_idx:
    for r_idx, row in enumerate(ws.iter_rows(min_row=header_idx+1, values_only=True), start=header_idx+1):
        vals = list(row)
        row_str = ' '.join(str(v) for v in vals if v)
        if '2641' in row_str or '6337' in row_str:
            print(f"\nV2 row (sheet row {r_idx}):")
            for ci, v in enumerate(vals):
                if v and ci in col_map:
                    print(f"  {col_map[ci]}: {v}")

print("\nDone.")
