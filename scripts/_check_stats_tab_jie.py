"""Check the Stats tab EV value from Joe's Sheet."""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

SHEET_URL = "https://docs.google.com/spreadsheets/d/1J-pZGelB9DxtahUc1JL3IXkT5C2_ajd_qvE_oqxUia4/edit"

# Fetch XLSX and read the Stats tab directly
import requests, io
from openpyxl import load_workbook

sheet_id = "1J-pZGelB9DxtahUc1JL3IXkT5C2_ajd_qvE_oqxUia4"
xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
resp = requests.get(xlsx_url, timeout=30)
wb = load_workbook(io.BytesIO(resp.content), data_only=True)

print("Sheet names:", wb.sheetnames)

if 'Stats' in wb.sheetnames:
    ws = wb['Stats']
    print("\n--- Stats Tab Contents (A:B, first 40 rows) ---")
    for row in ws.iter_rows(min_row=1, max_row=40, max_col=3, values_only=True):
        label = str(row[0]).strip() if row[0] else ''
        val = row[1] if len(row) > 1 else None
        val2 = row[2] if len(row) > 2 else None
        print(f"  {label:35s} | {val} | {val2}")
else:
    print("No 'Stats' tab found!")
    print("Available tabs:", wb.sheetnames)

wb.close()
