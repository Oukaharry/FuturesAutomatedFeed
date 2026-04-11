"""Check if any hedge result cells have comments (which might indicate formulas or notes)
and check for cells where XLSX value differs from CSV value."""
import sys, os, json, io, re
import requests
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

sheet_key = '1q4atojmjW03XLU6bRfubZ3WZiK071x3eQttt5kdKVYs'

# Fetch XLSX for raw cell values
print("Fetching XLSX...")
xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=xlsx"
resp = requests.get(xlsx_url, timeout=60)
wb = openpyxl.load_workbook(filename=io.BytesIO(resp.content), data_only=True)
ws = wb[wb.sheetnames[0]]

# Find header
header_idx = None
col_map = {}
for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
    row_vals = [str(c.value).strip() if c.value else '' for c in row]
    if any('Prop Firm' in v for v in row_vals):
        header_idx = r_idx
        for ci, h in enumerate(row_vals):
            if h: col_map[ci] = h
        break

print(f"Header at row {header_idx}")

# Find column indices for hedge results and statuses
target_cols = ['Status P1', 'Status',
    'Hedge Result 1', 'Hedge Result 2', 'Hedge Result 3', 'Hedge Result 4', 'Hedge Result 5',
    'Hedge Result 1.1', 'Hedge Result 2.1', 'Hedge Result 3.1', 'Hedge Result 4.1', 'Hedge Result 5.1',
    'Hedge Result 6', 'Hedge Result 7']
col_idx = {}
for ci, name in col_map.items():
    if name in target_cols:
        col_idx[name] = ci

print(f"Column indices: {col_idx}")

# Also fetch CSV for comparison
from utils.data_processor import fetch_evaluations
print("Fetching CSV...")
csv_evals, _ = fetch_evaluations(f"https://docs.google.com/spreadsheets/d/{sheet_key}/edit")
print(f"CSV rows: {len(csv_evals)}")

def pc(val):
    if val is None: return 0.0
    s = str(val).strip()
    if not s or s == '-' or s.lower() == 'none': return 0.0
    s = s.replace('$', '').replace(',', '').strip()
    try: return float(s)
    except: return 0.0

P1_COLS = ['Hedge Result 1', 'Hedge Result 2', 'Hedge Result 3', 'Hedge Result 4', 'Hedge Result 5']
FD_COLS = ['Hedge Result 1.1', 'Hedge Result 2.1', 'Hedge Result 3.1', 'Hedge Result 4.1', 'Hedge Result 5.1', 'Hedge Result 6', 'Hedge Result 7']

# Read XLSX data rows and compare with CSV
print("\n=== XLSX vs CSV differences in hedge cells (completed rows only) ===")
csv_idx = 0
xlsx_total = 0.0
csv_total = 0.0

for row_cells in ws.iter_rows(min_row=header_idx+1, values_only=False):
    prop_firm_ci = col_idx.get('Status P1', 0)
    # Check if row has Prop Firm value (column 0)
    pf_val = row_cells[0].value
    if not pf_val or not str(pf_val).strip():
        continue
    
    if csv_idx >= len(csv_evals):
        break
    
    csv_ev = csv_evals[csv_idx]
    rn = csv_idx + 2
    
    # Get statuses
    p1_status_ci = col_idx.get('Status P1')
    f_status_ci = col_idx.get('Status')
    p1_status = str(row_cells[p1_status_ci].value or '').strip() if p1_status_ci is not None else ''
    f_status = str(row_cells[f_status_ci].value or '').strip() if f_status_ci is not None else ''
    
    # Check completed rows
    is_p1_fail = p1_status == 'Fail'
    is_funded_ended = f_status in ('Fail', 'Completed')
    
    if is_p1_fail or is_funded_ended:
        for cname in P1_COLS + FD_COLS:
            ci = col_idx.get(cname)
            if ci is None:
                continue
            
            xlsx_val = row_cells[ci].value
            xlsx_num = pc(xlsx_val)
            csv_num = pc(csv_ev.get(cname))
            
            # Determine if this cell contributes to completed hedging
            contributes = False
            if cname in P1_COLS and is_p1_fail:
                contributes = True
            if cname in P1_COLS and is_funded_ended:
                contributes = True
            if cname in FD_COLS and is_funded_ended:
                contributes = True
            
            if contributes and abs(xlsx_num - csv_num) > 0.005:
                print(f"  Row {rn}: {pf_val} | {cname}: XLSX={xlsx_val}({xlsx_num:.2f}) CSV={csv_ev.get(cname)}({csv_num:.2f}) diff={xlsx_num-csv_num:.2f}  P1={p1_status} F={f_status}")
            
            if contributes:
                xlsx_total += xlsx_num
                csv_total += csv_num
    
    csv_idx += 1

print(f"\nXLSX total completed hedging: {xlsx_total:.2f}")
print(f"CSV total completed hedging:  {csv_total:.2f}")
print(f"Difference:                   {xlsx_total - csv_total:.2f}")
print(f"Sheet Stats tab says:         -439.52")
