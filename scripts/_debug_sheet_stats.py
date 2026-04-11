"""Fetch the Stats tab from Gregory Falk's Google Sheet to compare Hedging Review values."""
import requests
import re
import io
import openpyxl

SHEET_URL = "https://docs.google.com/spreadsheets/d/1in4Z-76-GJ2URCslKafg-RIsY3XNqRLzhcZQKjgFZuQ/edit?usp=sharing"

# Extract sheet key
match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', SHEET_URL)
sheet_key = match.group(1)
print(f"Sheet key: {sheet_key}")

# Fetch XLSX
xlsx_url = f"https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=xlsx"
print(f"Fetching XLSX from: {xlsx_url}")
resp = requests.get(xlsx_url, timeout=30)
print(f"Status: {resp.status_code}, Size: {len(resp.content)} bytes")

if resp.status_code == 200 and len(resp.content) > 1000:
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    print(f"\nSheet names: {wb.sheetnames}")
    
    # Find Stats tab
    stats_name = None
    for name in wb.sheetnames:
        if 'stat' in name.lower():
            stats_name = name
            break
    
    if stats_name:
        ws = wb[stats_name]
        print(f"\n=== Stats Tab: '{stats_name}' (rows: {ws.max_row}, cols: {ws.max_column}) ===\n")
        
        # Print ALL rows to see the full structure
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), max_col=min(ws.max_column, 5), values_only=False):
            row_num = row[0].row
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    vals.append(f"{cell.column_letter}{row_num}={v}")
            if vals:
                print(f"  Row {row_num:2d}: {' | '.join(vals)}")
    else:
        print("No Stats tab found!")
        # Print all tabs and their first few rows
        for name in wb.sheetnames:
            ws = wb[name]
            print(f"\n--- Tab: {name} ---")
            for row in ws.iter_rows(min_row=1, max_row=3, max_col=5, values_only=True):
                print(f"  {row}")
else:
    print(f"Failed to fetch XLSX. Trying CSV approach...")
    # Try fetching the first sheet as CSV
    csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_key}/export?format=csv"
    resp2 = requests.get(csv_url, timeout=30)
    print(f"CSV Status: {resp2.status_code}")
    if resp2.status_code == 200:
        lines = resp2.text.split('\n')
        for i, line in enumerate(lines[:30]):
            print(f"  {i}: {line[:200]}")
